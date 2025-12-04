import io
import openpyxl
from datetime import datetime
from fastapi import UploadFile

from app.repositories.whitelist_repository import WhitelistRepository
from app.models.whitelist import Whitelist
from app.models.admin import Admin

class WhitelistService:
    def __init__(self, repository: WhitelistRepository, current_user: Admin):
        self.repository = repository
        self.current_user = current_user
        print(current_user)

    def get_by_wallet_address(self, wallet_address: str):
        whitelist = self.repository.get_by_wallet_address(wallet_address)
        return whitelist
    
    async def import_whitelist(self, file: UploadFile):
        contents = await file.read()

        if not contents:
            raise ValueError("No file uploaded or file is empty")

        try:
            workbook = openpyxl.load_workbook(io.BytesIO(contents))
        except Exception as e:
            raise ValueError(f"Invalid Excel file: {str(e)}")
        
        whitelists = []
        for sheet in workbook.worksheets:
            for idx, row in enumerate(sheet.iter_rows(min_row=1, values_only=True), start=1):
                wallet_address, whitelist_type = row[0], row[1]

                if not row[1]:
                    print(f"Null found at row {idx}: {row}")
                    continue

                whitelists.append({
                    "wallet_address": wallet_address.strip() if isinstance(wallet_address, str) else wallet_address,
                    "whitelist_type": whitelist_type.strip() if isinstance(whitelist_type, str) else whitelist_type,
                    "created_by": self.current_user['id'],
                    "created_at": datetime.now()
                })

        self.repository.add(whitelists)
        # print(whitelists);
        return len(whitelists)

    def export_to_excel(self):
        records = self.repository.get_all()

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Wallet Whitelist'

        sheet.append(['No', 'Wallet Address', 'Whitelist Type'])

        for idx, row in enumerate(records, start=1):
            sheet.append([idx, row.wallet_address, row.whitelist_type])

        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        return buffer